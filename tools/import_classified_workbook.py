"""Convert the authoritative classified workbook into the JSON master seed."""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def value(cell):
    raw = cell.value
    if raw is None or isinstance(raw, (str, int, float, bool)):
        return raw
    return str(raw)


def main(source: Path, target: Path) -> None:
    workbook = load_workbook(source, data_only=True, read_only=True)
    sheets = {}
    for name in ("Basis Data Indikator", "Data Target-Realisasi"):
        sheet = workbook[name]
        sheets[name] = [[value(cell) for cell in row] for row in sheet.iter_rows()]
    target.write_text(
        json.dumps({"source": source.name, "sheets": sheets}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main(Path(sys.argv[1]), Path(sys.argv[2]))
