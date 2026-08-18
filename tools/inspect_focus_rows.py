from pathlib import Path

from openpyxl import load_workbook

source = Path(r"C:\Users\user\Downloads\BASIS_DATA_INDIKATOR_ISV-IUP_KALTARA_terklasifikasi.xlsx")
book = load_workbook(source, data_only=True, read_only=True)
for sheet_name in ("Basis Data Indikator", "Data Target-Realisasi"):
    sheet = book[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    print("\n", sheet_name, headers)
    for values in rows:
        row = dict(zip(headers, values, strict=False))
        if str(row.get("ID Indikator", "")).strip() in {"ISV-001", "ISV-004", "ISV-005", "IUP-028", "IUP-050"}:
            print(row)
