"""Unggahan Excel massal: staging, diff, dan penerapan setelah disetujui.

Alurnya sengaja dua langkah. Berkas yang diunggah dijalankan lewat pipeline ETL
ke basis data staging terpisah, hasilnya dibandingkan dengan data berjalan, dan
baru diterapkan setelah admin menyetujui diff-nya.
"""

from __future__ import annotations

import json
import sqlite3
import uuid
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, NamedTuple

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..config import settings
from ..models import KODE_PROVINSI, JenisNilai, StatusUnggahan, StatusVerifikasi
from ..repositories import nilai as repo_nilai
from ..repositories import tata_kelola as repo_tata_kelola


def sheet_wajib() -> frozenset[str]:
    """Sheet yang harus ada agar pipeline ETL dapat berjalan.

    Diambil dari konfigurasi ETL, bukan disalin, supaya daftar ini tidak
    menyimpang ketika versi workbook berganti.
    """
    from src.etl.config import bawaan

    return frozenset(bawaan().sheet_wajib)


class BerkasTidakValid(Exception):
    """Berkas unggahan tidak memenuhi syarat untuk diproses."""


class HasilPratinjau(NamedTuple):
    path_arsip: Path
    path_staging: Path
    diff: dict[str, Any]


def berekstensi_xlsx(nama_berkas: str | None) -> bool:
    return bool(nama_berkas) and str(nama_berkas).lower().endswith(".xlsx")


def ukuran_wajar(jumlah_byte: int) -> bool:
    return jumlah_byte <= settings.max_unggah_bytes


def arsipkan(isi: bytes) -> Path:
    direktori = Path(settings.archive_dir)
    direktori.mkdir(parents=True, exist_ok=True)
    berkas = direktori / f"{datetime.now(UTC):%Y%m%d-%H%M%S}-{uuid.uuid4()}.xlsx"
    berkas.write_bytes(isi)
    return berkas


def periksa_sheet(path: Path) -> None:
    workbook = load_workbook(path, read_only=True)
    try:
        hilang = sheet_wajib() - set(workbook.sheetnames)
    finally:
        workbook.close()
    if hilang:
        raise BerkasTidakValid(f"Sheet hilang: {', '.join(sorted(hilang))}")


def _baca_staging(path: Path) -> tuple[dict[str, str], dict[tuple[str, int, str], float | None]]:
    koneksi = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
    try:
        indikator = {
            baris[0]: baris[1] for baris in koneksi.execute("SELECT id_indikator,nama_indikator FROM indikator")
        }
        nilai = {
            (baris[0], baris[1], baris[2]): baris[3]
            for baris in koneksi.execute("SELECT id_indikator,tahun,jenis,nilai FROM nilai_indikator")
        }
    finally:
        koneksi.close()
    return indikator, nilai


def susun_diff(
    session: Session, path_staging: Path
) -> tuple[dict[str, Any], dict[tuple[str, int, str], tuple[float | None, str | None]]]:
    """Bandingkan hasil ETL staging dengan nilai provinsi yang berlaku."""
    indikator_baru, nilai_baru = _baca_staging(path_staging)
    indikator_lama = {item.id_indikator: item.nama_indikator for item in repo_nilai.semua_indikator_ringkas(session)}
    nilai_lama = {
        (baris.id_indikator, baris.tahun, baris.jenis): baris.nilai
        for baris in repo_nilai.semua_nilai_provinsi(session)
    }
    diff = {
        "indikator_baru": sorted(set(indikator_baru) - set(indikator_lama)),
        "indikator_hilang": sorted(set(indikator_lama) - set(indikator_baru)),
        "nilai_berubah": [
            {
                "id": kunci[0],
                "tahun": kunci[1],
                "jenis": kunci[2],
                "lama": nilai_lama.get(kunci),
                "baru": nilai_baru.get(kunci),
            }
            for kunci in sorted(set(nilai_lama) | set(nilai_baru))
            if nilai_lama.get(kunci) != nilai_baru.get(kunci)
        ],
    }
    return diff, {}


def terapkan(session: Session, unggahan: Any, pengguna_id: int | None) -> int:
    """Tulis nilai dari staging ke tabel fakta; kembalikan jumlah perubahan.

    Hanya indikator yang sudah dikenal yang disentuh: unggahan tidak boleh
    diam-diam membuat dimensi indikator baru.
    """
    path_staging = Path(unggahan.path_arsip).with_suffix(".stage.db")
    if not path_staging.exists():
        raise BerkasTidakValid("Database staging tidak ditemukan")

    _, nilai_baru = _baca_staging(path_staging)
    dikenal = {item.id_indikator for item in repo_nilai.semua_indikator_ringkas(session)}
    jumlah = 0
    for (id_indikator, tahun, jenis), nilai in nilai_baru.items():
        if id_indikator not in dikenal or jenis not in tuple(JenisNilai):
            continue
        baris_lama = repo_nilai.ambil(session, id_indikator, KODE_PROVINSI, tahun, jenis)
        if baris_lama is not None and baris_lama.nilai == nilai:
            continue
        _, lama = repo_nilai.upsert(
            session,
            id_indikator=id_indikator,
            wilayah_kode=KODE_PROVINSI,
            tahun=tahun,
            jenis=jenis,
            nilai=nilai,
            nilai_teks=baris_lama.nilai_teks if baris_lama else None,
            satuan_catatan=baris_lama.satuan_catatan if baris_lama else None,
            sumber=unggahan.nama_file_asli,
            status_verifikasi=StatusVerifikasi.DISETUJUI,
        )
        repo_tata_kelola.catat_perubahan(
            session,
            pengguna_id=pengguna_id,
            id_indikator=id_indikator,
            field=f"nilai:{tahun}:{jenis}",
            nilai_lama=None if lama is None else str(lama),
            nilai_baru=str(nilai),
            sumber_perubahan="unggah",
            referensi_id=str(unggahan.id),
        )
        jumlah += 1

    unggahan.status = StatusUnggahan.DISETUJUI
    unggahan.disetujui_pada = datetime.now(UTC)
    return jumlah


def ringkasan_diff_json(diff: dict[str, Any]) -> str:
    return json.dumps(diff, ensure_ascii=False)
