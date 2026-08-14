from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from hashlib import sha256
from io import BytesIO, StringIO
import csv, json, math, os, re, shutil, sqlite3, tempfile, uuid, zipfile
from pathlib import Path

import jwt
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, Response, StreamingResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from openpyxl import Workbook, load_workbook
from pwdlib import PasswordHash
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy import text
from sqlalchemy.orm import Session

from .database import DEFAULT_DB, get_db
from src.etl.pipeline import run as run_etl


router = APIRouter(prefix="/api/v1")
oauth2 = OAuth2PasswordBearer(tokenUrl="/api/v1/auth/login", auto_error=False)
password_hash = PasswordHash.recommended()
SECRET = os.getenv("SEBATIK_SECRET_KEY", "GANTI-SECRET-INI-SEBELUM-PRODUKSI-SEBATIK")
ARCHIVE_DIR = Path(os.getenv("SEBATIK_ARCHIVE_DIR", DEFAULT_DB.parent / "arsip-unggahan"))
EVIDENCE_DIR = Path(os.getenv("SEBATIK_EVIDENCE_DIR", DEFAULT_DB.parent / "bukti-dukung"))

# Lima indikator sorotan beranda. Sejak kartu makro berjalan sebagai korsel,
# daftar ini tidak lagi membatasi apa yang tampil — ia hanya menentukan urutan
# tampil pertama. Sisanya diambil dari klasifikasi kelompok_makro di basis data
# dan menyusul di belakangnya.
HOME_MACRO_IDS=("ISV-001","IUP-050","ISV-004","ISV-005","IUP-028")


def home_macro_ids(db):
    """Seluruh indikator berklasifikasi makro, lima sorotan lebih dulu."""
    classified=[r["id_indikator"] for r in rows(db,"SELECT id_indikator FROM beranda_indikator WHERE kelompok_makro LIKE 'Makro%' AND status_verifikasi='DISETUJUI' ORDER BY kategori DESC,id_indikator")]
    highlighted=[iid for iid in HOME_MACRO_IDS if iid in classified]
    return highlighted+[iid for iid in classified if iid not in highlighted]


def latest_period_value(db,indicator_id,year,wilayah_kode="65"):
    if wilayah_kode=="65":
        return one(db,"SELECT nilai,periode,label_periode FROM beranda_nilai_periode WHERE id_indikator=:id AND tahun=:year AND jenis='realisasi' AND status_verifikasi='DISETUJUI' ORDER BY periode DESC LIMIT 1",{"id":indicator_id,"year":year})
    return one(db,"SELECT nilai,periode,label_periode FROM beranda_nilai_wilayah_periode WHERE id_indikator=:id AND wilayah_kode=:wilayah AND tahun=:year AND jenis='realisasi' AND status_verifikasi='DISETUJUI' ORDER BY periode DESC LIMIT 1",{"id":indicator_id,"year":year,"wilayah":wilayah_kode})


def rows(db, sql, params=None): return [dict(r._mapping) for r in db.execute(text(sql), params or {}).all()]
def one(db, sql, params=None):
    result=db.execute(text(sql), params or {}).first(); return dict(result._mapping) if result else None


def current_user(token: str | None=Depends(oauth2), db: Session=Depends(get_db)):
    if not token: return {"id":None,"username":"pengunjung","nama":"Pengunjung","peran":"PENGUNJUNG","tim_pjk":None,"wilayah_kode":None}
    try: payload=jwt.decode(token,SECRET,algorithms=["HS256"])
    except jwt.PyJWTError: raise HTTPException(401,"Token tidak valid")
    user=one(db,"SELECT id,username,nama,peran,tim_pjk,wilayah_kode,harus_ganti_password FROM pengguna WHERE id=:id AND aktif=1",{"id":int(payload["sub"])})
    if not user: raise HTTPException(401,"Pengguna tidak aktif")
    return user


def require(*roles):
    def dependency(user=Depends(current_user)):
        if user["peran"] not in roles: raise HTTPException(403,"Akses tidak diizinkan")
        return user
    return dependency


def achievement(actual, target, direction, verified):
    if actual is None or target is None or not direction or not verified: return None,"BELUM_ADA_DATA"
    if direction=="NAIK": pct=(actual/target*100) if target else None
    else: pct=(target/actual*100) if actual else None
    if pct is None: return None,"BELUM_ADA_DATA"
    return round(pct,2), "TERCAPAI" if pct>=100 else "MENDEKATI" if pct>=90 else "PERLU_PERHATIAN"


def _last_number(value, text_value=None):
    if value is not None: return float(value)
    if text_value:
        matches=re.findall(r"-?\d+(?:[.,]\d+)?",str(text_value))
        try: return float(matches[-1].replace(",", ".")) if matches else None
        except ValueError: return None
    return None


def availability_dimensions(db:Session):
    """Kelengkapan slot realisasi 2021-2025 pada empat lapis klasifikasi."""
    dimensions=(
        ("sasaran_visi","Sasaran Visi",5),
        ("misi_agenda","Misi/Agenda Pembangunan",8),
        ("arah_ie","Arah Pembangunan",17),
        ("indikator_induk","Indikator Utama Induk",45),
    )
    result=[]
    for column,label,total_groups in dimensions:
        indicators=rows(db,f"SELECT id_indikator,{column} kelompok FROM beranda_indikator WHERE status_verifikasi='DISETUJUI' AND {column} IS NOT NULL AND trim({column})<>'' AND {column} NOT LIKE '-%'")
        ids=[x["id_indikator"] for x in indicators]
        filled=0
        if ids:
            placeholders=",".join(f":i{i}" for i in range(len(ids)))
            filled=db.scalar(text(f"SELECT COUNT(*) FROM beranda_nilai WHERE id_indikator IN ({placeholders}) AND jenis='realisasi' AND tahun BETWEEN 2021 AND 2025 AND status_verifikasi='DISETUJUI' AND (nilai IS NOT NULL OR nilai_teks IS NOT NULL)"),{f"i{i}":v for i,v in enumerate(ids)}) or 0
        possible=len(ids)*5
        result.append({"kode":column,"label":label,"jumlah_kelompok":total_groups,"jumlah_indikator":len(ids),"slot_terisi":filled,"slot_total":possible,"persentase":round(filled/possible*100,1) if possible else 0})
    return result


@router.get("/beranda")
def homepage(tahun:int|None=None,wilayah_kode:str="65",db:Session=Depends(get_db)):
    if not one(db,"SELECT 1 ok FROM wilayah WHERE kode=:k AND aktif=1",{"k":wilayah_kode}):raise HTTPException(422,"Wilayah tidak valid")
    years=[r[0] for r in db.execute(text("SELECT DISTINCT tahun FROM beranda_nilai WHERE jenis='realisasi' AND status_verifikasi='DISETUJUI' ORDER BY tahun")).all()]
    if not years: return {"tahun":tahun,"tahun_tersedia":[],"indikator_makro":[],"sasaran_visi":[],"ketersediaan_kelompok":availability_dimensions(db)}
    selected=tahun if tahun in years else max(years)
    source_table="beranda_nilai" if wilayah_kode=="65" else "beranda_nilai_wilayah"
    region_filter="" if wilayah_kode=="65" else " AND wilayah_kode=:wilayah"
    params={"wilayah":wilayah_kode}

    macro_ids=home_macro_ids(db)
    macros=[]
    for iid in macro_ids:
        indicator=one(db,"SELECT id_indikator,nama_indikator,arah_pembangunan,kode_indikator,satuan FROM beranda_indikator WHERE id_indikator=:id AND status_verifikasi='DISETUJUI'",{"id":iid})
        if not indicator: continue
        current=one(db,f"SELECT nilai,nilai_teks,satuan_catatan FROM {source_table} WHERE id_indikator=:id AND tahun=:year AND jenis='realisasi' AND status_verifikasi='DISETUJUI'{region_filter}",{"id":iid,"year":selected,**params})
        previous=one(db,f"SELECT nilai,nilai_teks FROM {source_table} WHERE id_indikator=:id AND tahun<:year AND jenis='realisasi' AND status_verifikasi='DISETUJUI'{region_filter} ORDER BY tahun DESC LIMIT 1",{"id":iid,"year":selected,**params})
        target=one(db,f"SELECT nilai,nilai_teks FROM {source_table} WHERE id_indikator=:id AND tahun=:year AND jenis='target' AND status_verifikasi='DISETUJUI'{region_filter}",{"id":iid,"year":selected,**params})
        period_current=latest_period_value(db,iid,selected,wilayah_kode)
        cv=period_current["nilai"] if period_current else _last_number(current.get("nilai") if current else None,current.get("nilai_teks") if current else None);pv=_last_number(previous.get("nilai") if previous else None,previous.get("nilai_teks") if previous else None)
        macros.append({**indicator,"tahun":selected,"nilai":period_current["nilai"] if period_current else current.get("nilai") if current else None,"nilai_teks":None if period_current else current.get("nilai_teks") if current else None,"target":target.get("nilai") if target else None,"target_teks":target.get("nilai_teks") if target else None,"perubahan":round(cv-pv,2) if cv is not None and pv is not None else None,"arah_perubahan":"NAIK" if cv is not None and pv is not None and cv>pv else "TURUN" if cv is not None and pv is not None and cv<pv else "TETAP" if cv is not None and pv is not None else "TIDAK_ADA_DATA","keterangan":period_current["label_periode"] if period_current else "Data belum tersedia pada tahun dipilih" if not current else current.get("satuan_catatan")})

    vision=[]
    for indicator in rows(db,"SELECT id_indikator,kode_indikator,nama_indikator,arah_pembangunan,satuan FROM beranda_indikator WHERE kategori='ISV' AND status_verifikasi='DISETUJUI' ORDER BY id_indikator"):
        actual=one(db,f"SELECT nilai,nilai_teks,satuan_catatan FROM {source_table} WHERE id_indikator=:id AND tahun=:year AND jenis='realisasi' AND status_verifikasi='DISETUJUI'{region_filter}",{"id":indicator["id_indikator"],"year":selected,**params})
        target=one(db,f"SELECT nilai,nilai_teks FROM {source_table} WHERE id_indikator=:id AND tahun=:year AND jenis='target' AND status_verifikasi='DISETUJUI'{region_filter}",{"id":indicator["id_indikator"],"year":selected,**params})
        vision.append({**indicator,"tahun":selected,"nilai":actual.get("nilai") if actual else None,"nilai_teks":actual.get("nilai_teks") if actual else None,"target":target.get("nilai") if target else None,"target_teks":target.get("nilai_teks") if target else None,"keterangan":"Data belum tersedia pada tahun dipilih" if not actual else actual.get("satuan_catatan")})
    return {"tahun":selected,"wilayah_kode":wilayah_kode,"tahun_tersedia":years,"indikator_makro":macros,"sasaran_visi":vision,"ketersediaan_kelompok":availability_dimensions(db),"status_data":"HANYA_TERVERIFIKASI"}


@router.get("/indikator-explorer")
def indicator_explorer(db:Session=Depends(get_db)):
    data=rows(db,"SELECT id_indikator,kategori,kelompok,arah_pembangunan,kode_indikator,nama_indikator,satuan FROM beranda_indikator WHERE status_verifikasi='DISETUJUI' ORDER BY kelompok,kategori DESC,id_indikator")
    groups=[]
    for name in dict.fromkeys(x["kelompok"] or "Tanpa Kelompok" for x in data):
        items=[x for x in data if (x["kelompok"] or "Tanpa Kelompok")==name]
        groups.append({"kelompok":name,"jumlah":len(items),"indikator":items})
    return {"data":groups,"total_indikator":len(data),"status_data":"HANYA_TERVERIFIKASI"}


@router.get("/indikator-explorer/{indicator_id}")
def indicator_explorer_detail(indicator_id:str,tahun:int|None=None,db:Session=Depends(get_db)):
    indicator=one(db,"SELECT id_indikator,kategori,kelompok,arah_pembangunan,kode_indikator,nama_indikator,satuan,sumber_data,frekuensi,opd_pengampu FROM beranda_indikator WHERE id_indikator=:id AND status_verifikasi='DISETUJUI'",{"id":indicator_id})
    if not indicator: raise HTTPException(404,"Indikator tidak ditemukan atau belum diverifikasi")
    raw=rows(db,"SELECT tahun,jenis,nilai,nilai_teks,satuan_catatan FROM beranda_nilai WHERE id_indikator=:id AND status_verifikasi='DISETUJUI' ORDER BY tahun,jenis",{"id":indicator_id})
    years=sorted({x["tahun"] for x in raw if x["jenis"]=="realisasi"})
    selected=tahun if tahun in years else (max(years) if years else None)
    actuals={x["tahun"]:x for x in raw if x["jenis"]=="realisasi"};targets={x["tahun"]:x for x in raw if x["jenis"]=="target"}
    timeline=[];previous=None
    for year in sorted(set(actuals)|set(targets)):
        actual=actuals.get(year);target=targets.get(year);numeric=_last_number(actual.get("nilai") if actual else None,actual.get("nilai_teks") if actual else None)
        growth=round((numeric-previous)/abs(previous)*100,2) if numeric is not None and previous not in (None,0) else None
        timeline.append({"tahun":year,"realisasi":actual.get("nilai") if actual else None,"realisasi_teks":actual.get("nilai_teks") if actual else None,"target":target.get("nilai") if target else None,"target_teks":target.get("nilai_teks") if target else None,"growth":growth})
        if numeric is not None: previous=numeric
    regions=rows(db,"SELECT kode,nama,tingkat FROM wilayah WHERE parent_kode='65' AND aktif=1 ORDER BY kode")
    for region in regions:
        value=one(db,"SELECT nilai,nilai_teks FROM beranda_nilai_wilayah WHERE id_indikator=:id AND wilayah_kode=:w AND tahun=:t AND jenis='realisasi' AND status_verifikasi='DISETUJUI'",{"id":indicator_id,"w":region["kode"],"t":selected}) if selected else None
        target=one(db,"SELECT nilai,nilai_teks FROM beranda_nilai_wilayah WHERE id_indikator=:id AND wilayah_kode=:w AND tahun=:t AND jenis='target' AND status_verifikasi='DISETUJUI'",{"id":indicator_id,"w":region["kode"],"t":selected}) if selected else None
        region.update({"tahun":selected,"nilai":value.get("nilai") if value else None,"nilai_teks":value.get("nilai_teks") if value else None,"target":target.get("nilai") if target else None,"target_teks":target.get("nilai_teks") if target else None,"status":"TERSEDIA" if value else "BELUM_ADA_DATA"})
    return {**indicator,"tahun":selected,"tahun_tersedia":years,"series":timeline,"wilayah":regions,"status_data":"HANYA_TERVERIFIKASI","catatan_wilayah":"Data kabupaten/kota belum tersedia. Nilai akan tampil setelah operator wilayah mengirim dan verifikator menyetujui."}


@router.get("/capaian-explorer")
def achievement_explorer(db:Session=Depends(get_db)):
    indicators=rows(db,"SELECT id_indikator,kategori,kelompok,arah_pembangunan,kode_indikator,nama_indikator,satuan FROM beranda_indikator WHERE status_verifikasi='DISETUJUI' ORDER BY kelompok,id_indikator")
    return {"indikator":indicators,"kelompok":sorted({x["kelompok"] for x in indicators if x["kelompok"]}),"wilayah":rows(db,"SELECT kode,nama,tingkat FROM wilayah WHERE aktif=1 ORDER BY length(kode),kode"),"status_data":"HANYA_TERVERIFIKASI"}


@router.get("/capaian-explorer/{indicator_id}")
def achievement_explorer_detail(indicator_id:str,tahun:int|None=None,wilayah_kode:str="65",db:Session=Depends(get_db)):
    indicator=one(db,"SELECT id_indikator,kategori,kelompok,arah_pembangunan,kode_indikator,nama_indikator,satuan,sumber_data,frekuensi,opd_pengampu FROM beranda_indikator WHERE id_indikator=:id AND status_verifikasi='DISETUJUI'",{"id":indicator_id})
    if not indicator: raise HTTPException(404,"Indikator tidak ditemukan atau belum diverifikasi")
    region=one(db,"SELECT kode,nama,tingkat FROM wilayah WHERE kode=:k AND aktif=1",{"k":wilayah_kode})
    if not region: raise HTTPException(422,"Wilayah tidak valid")
    source="beranda_nilai" if wilayah_kode=="65" else "beranda_nilai_wilayah"
    region_sql="" if wilayah_kode=="65" else " AND wilayah_kode=:wilayah"
    params={"id":indicator_id,"wilayah":wilayah_kode}
    raw=rows(db,f"SELECT tahun,jenis,nilai,nilai_teks FROM {source} WHERE id_indikator=:id AND status_verifikasi='DISETUJUI'{region_sql} ORDER BY tahun,jenis",params)
    actual_rows=[x for x in raw if x["jenis"]=="realisasi" and _last_number(x["nilai"],x["nilai_teks"]) is not None]
    available_years=[x["tahun"] for x in actual_rows]
    selected=tahun if tahun in available_years else (max(available_years) if available_years else tahun)
    targets={x["tahun"]:x for x in raw if x["jenis"]=="target"}
    target_2045=targets.get(2045);target_num=_last_number(target_2045["nilai"],target_2045["nilai_teks"]) if target_2045 else None
    # Target antara RPJMD. Dipakai sebagai tolok progres pada tracker karena
    # 2029 adalah horizon yang masih bisa ditindaklanjuti perencana hari ini;
    # 2045 tetap dikirim sebagai tujuan akhir.
    target_2029=targets.get(2029);target_2029_num=_last_number(target_2029["nilai"],target_2029["nilai_teks"]) if target_2029 else None
    series=[];previous=None
    for item in actual_rows:
        numeric=_last_number(item["nilai"],item["nilai_teks"])
        growth=round((numeric-previous)/abs(previous)*100,2) if previous not in (None,0) else None
        series.append({"tahun":item["tahun"],"nilai":numeric,"nilai_asli":item["nilai"],"nilai_teks":item["nilai_teks"],"growth":growth,"target":_last_number(targets[item["tahun"]]["nilai"],targets[item["tahun"]]["nilai_teks"]) if item["tahun"] in targets else None})
        previous=numeric
    current=next((x for x in series if x["tahun"]==selected),None);baseline=series[0] if series else None;previous_item=next((x for x in reversed(series) if selected and x["tahun"]<selected),None)
    direction="NAIK" if baseline and target_num is not None and target_num>=baseline["nilai"] else "TURUN" if baseline and target_num is not None else None
    def progress_towards(goal):
        """Bagian jalan yang sudah ditempuh dari baseline menuju satu target.

        Dijepit ke 0–100 supaya cincin tracker tidak pernah tergambar melebihi
        lingkaran penuh atau berbalik arah ketika realisasi melewati target
        atau bergerak menjauh dari baseline."""
        if not (current and baseline and goal is not None): return None
        denominator=goal-baseline["nilai"]
        if denominator==0: return 100 if current["nilai"]==goal else None
        return max(0,min(100,round((current["nilai"]-baseline["nilai"])/denominator*100,1)))

    progress=progress_towards(target_num)
    progress_2029=progress_towards(target_2029_num)
    gap=round(target_num-current["nilai"],4) if current and target_num is not None else None
    gap_2029=round(target_2029_num-current["nilai"],4) if current and target_2029_num is not None else None
    annual=round(gap/(2045-selected),4) if gap is not None and selected and selected<2045 else None
    improving=None
    if current and previous_item and direction: improving=current["nilai"]>=previous_item["nilai"] if direction=="NAIK" else current["nilai"]<=previous_item["nilai"]
    if not current: insight=f"Data realisasi terverifikasi untuk {region['nama']} pada tahun yang dipilih belum tersedia."
    elif target_2029_num is None and target_num is None: insight="Target 2029 dan 2045 belum tersedia sehingga progres belum dapat dihitung."
    else:
        trend="membaik" if improving is True else "menjauh dari arah target" if improving is False else "belum dapat dibandingkan dengan tahun sebelumnya"
        # Kalimat ini mengikuti apa yang digambar cincin tracker — 2029 — supaya
        # angka yang dibaca dan angka yang dilihat tidak bercerita beda.
        if progress_2029 is not None:
            insight=f"Capaian {selected} berada pada {progress_2029}% perjalanan dari baseline {baseline['tahun']} menuju target 2029 dan trennya {trend}."
            if target_num is not None: insight+=f" Target akhir 2045 berada di {target_num:g}."
        elif progress is not None:
            insight=f"Target 2029 belum tersedia. Terhadap target akhir 2045, capaian {selected} berada pada {progress}% perjalanan dari baseline {baseline['tahun']} dan trennya {trend}."
        else:
            insight=f"Capaian {selected} tersedia, tetapi progres belum dapat dihitung lengkap."
    projection=[{"tahun":x["tahun"],"realisasi":x["nilai"],"jalur_target":current["nilai"] if current and x["tahun"]==selected else None} for x in series]
    if target_num is not None and not any(x["tahun"]==2045 for x in projection):projection.append({"tahun":2045,"realisasi":None,"jalur_target":target_num})
    return {**indicator,"wilayah":region,"tahun":selected,"tahun_tersedia":available_years,"series":series,"projection":projection,"nilai_tahun":current["nilai"] if current else None,"nilai_teks":current["nilai_teks"] if current else None,"target_2045":target_num,"target_2045_teks":target_2045["nilai_teks"] if target_2045 else None,"target_2029":target_2029_num,"target_2029_teks":target_2029["nilai_teks"] if target_2029 else None,"arah_target":direction,"progres_2045":progress,"progres_2029":progress_2029,"gap_2045":gap,"gap_2029":gap_2029,"kebutuhan_per_tahun":annual,"insight":insight,"status_data":"HANYA_TERVERIFIKASI","catatan_wilayah":None if wilayah_kode=="65" else "Belum ada basis data kabupaten/kota. Visualisasi akan terisi setelah data wilayah diverifikasi."}


@router.get("/insight")
def insight_dashboard(indikator_id:str|None=None,wilayah_kode:str="65",db:Session=Depends(get_db)):
    region=one(db,"SELECT kode,nama,tingkat FROM wilayah WHERE kode=:k AND aktif=1",{"k":wilayah_kode})
    if not region: raise HTTPException(422,"Wilayah tidak valid")
    # Tanpa LIMIT: pemilih kartu di halaman Insight kini berupa rel mendatar,
    # jadi seluruh indikator berklasifikasi makro muat tanpa memotong daftar.
    # Urutannya tidak diubah supaya kartu pertama — yang otomatis terpilih saat
    # halaman dibuka — tetap indikator yang sama seperti sebelumnya.
    indicators=rows(db,"SELECT id_indikator,kode_indikator,nama_indikator,kelompok,satuan,sumber_data,opd_pengampu FROM beranda_indikator WHERE kelompok_makro LIKE 'Makro%' AND status_verifikasi='DISETUJUI' ORDER BY kategori DESC,id_indikator")
    source="beranda_nilai" if wilayah_kode=="65" else "beranda_nilai_wilayah";region_sql="" if wilayah_kode=="65" else " AND wilayah_kode=:wilayah"
    cards=[]
    for item in indicators:
        latest=one(db,f"SELECT tahun,nilai,nilai_teks FROM {source} WHERE id_indikator=:id AND jenis='realisasi' AND tahun<=:now AND status_verifikasi='DISETUJUI'{region_sql} AND (nilai IS NOT NULL OR nilai_teks IS NOT NULL) ORDER BY tahun DESC LIMIT 1",{"id":item["id_indikator"],"now":date.today().year,"wilayah":wilayah_kode})
        previous=one(db,f"SELECT tahun,nilai,nilai_teks FROM {source} WHERE id_indikator=:id AND jenis='realisasi' AND tahun<:year AND status_verifikasi='DISETUJUI'{region_sql} AND (nilai IS NOT NULL OR nilai_teks IS NOT NULL) ORDER BY tahun DESC LIMIT 1",{"id":item["id_indikator"],"year":latest["tahun"] if latest else date.today().year,"wilayah":wilayah_kode}) if latest else None
        # Sebagian indikator dilaporkan per semester atau triwulan. Bila untuk
        # tahun terakhir ada rilis periode yang sudah disetujui, angka itulah
        # yang paling mutakhir — bukan angka tahunannya, yang baru terisi
        # setelah setahun penuh lewat. Beranda sudah memakai aturan ini; kartu
        # Insight sebelumnya belum, sehingga keduanya bisa memperlihatkan angka
        # berbeda untuk indikator yang sama.
        period=latest_period_value(db,item["id_indikator"],latest["tahun"],wilayah_kode) if latest else None
        current_num=(period["nilai"] if period else None) or _last_number(latest["nilai"],latest["nilai_teks"]) if latest else None
        previous_num=_last_number(previous["nilai"],previous["nilai_teks"]) if previous else None
        # Label dirangkai lengkap dengan tahunnya: "Semester 2" saja tidak
        # memberi tahu semester tahun berapa.
        period_label=f"{period['label_periode']} {latest['tahun']}" if period and period.get("label_periode") else (str(latest["tahun"]) if latest else None)
        cards.append({**item,"tahun":latest["tahun"] if latest else None,"label_periode":period_label,"nilai":period["nilai"] if period else latest["nilai"] if latest else None,"nilai_teks":None if period else latest["nilai_teks"] if latest else None,"perubahan":round(current_num-previous_num,2) if current_num is not None and previous_num is not None else None,"status":"TERSEDIA" if latest else "BELUM_ADA_DATA"})
    selected=indikator_id if indikator_id in {x["id_indikator"] for x in cards} else (cards[0]["id_indikator"] if cards else None)
    detail=next((x for x in cards if x["id_indikator"]==selected),None)
    raw=rows(db,f"SELECT tahun,nilai,nilai_teks FROM {source} WHERE id_indikator=:id AND jenis='realisasi' AND status_verifikasi='DISETUJUI'{region_sql} ORDER BY tahun",{"id":selected,"wilayah":wilayah_kode}) if selected else []
    series=[];previous=None
    for value in raw:
        numeric=_last_number(value["nilai"],value["nilai_teks"])
        if numeric is None: continue
        growth=round((numeric-previous)/abs(previous)*100,2) if previous not in (None,0) else None
        series.append({"tahun":value["tahun"],"nilai":numeric,"nilai_teks":value["nilai_teks"],"growth":growth});previous=numeric
    selected_year=detail["tahun"] if detail else None
    comparisons=rows(db,"SELECT kode,nama,tingkat FROM wilayah WHERE parent_kode='65' AND aktif=1 ORDER BY kode")
    for reg in comparisons:
        value=one(db,"SELECT nilai,nilai_teks FROM beranda_nilai_wilayah WHERE id_indikator=:id AND wilayah_kode=:w AND tahun=:t AND jenis='realisasi' AND status_verifikasi='DISETUJUI'",{"id":selected,"w":reg["kode"],"t":selected_year}) if selected_year else None
        reg.update({"nilai":value["nilai"] if value else None,"nilai_teks":value["nilai_teks"] if value else None,"status":"TERSEDIA" if value else "BELUM_ADA_DATA"})
    return {"tahun_sistem":date.today().year,"wilayah":region,"wilayah_opsi":rows(db,"SELECT kode,nama,tingkat FROM wilayah WHERE aktif=1 ORDER BY length(kode),kode"),"indikator_makro":cards,"indikator_aktif":detail,"series":series,"perbandingan_wilayah":comparisons,"status_data":"HANYA_TERVERIFIKASI","catatan_wilayah":"Data kabupaten/kota belum tersedia. Peta dan bar chart akan terisi setelah data operator wilayah diverifikasi." if not any(x["status"]=="TERSEDIA" for x in comparisons) else None}


@router.get("/validitas")
def validity_table(wilayah_kode:str="65",q:str|None=None,user=Depends(current_user),db:Session=Depends(get_db)):
    region=one(db,"SELECT kode,nama,tingkat FROM wilayah WHERE kode=:k AND aktif=1",{"k":wilayah_kode})
    if not region: raise HTTPException(422,"Wilayah tidak valid")
    indicators=rows(db,"SELECT id_indikator,kode_indikator,nama_indikator,satuan,opd_pengampu,is_proxy,diverifikasi_pada FROM beranda_indikator WHERE status_verifikasi='DISETUJUI' ORDER BY id_indikator")
    if q: indicators=[x for x in indicators if q.lower() in x["nama_indikator"].lower() or q.lower() in x["kode_indikator"].lower()]
    data=[]
    for item in indicators:
        submission=one(db,"SELECT u.id usulan_id,n.diverifikasi_pada,u.dibuat_pada,p.nama pengusul,p.peran FROM beranda_nilai_wilayah n LEFT JOIN usulan_nilai u ON u.id=n.usulan_id LEFT JOIN pengguna p ON p.id=u.pengusul_id WHERE n.id_indikator=:id AND n.wilayah_kode=:w AND n.status_verifikasi='DISETUJUI' ORDER BY n.diverifikasi_pada DESC LIMIT 1",{"id":item["id_indikator"],"w":wilayah_kode})
        if wilayah_kode=="65":
            latest=one(db,"SELECT MAX(diverifikasi_pada) waktu,COUNT(*) jumlah FROM beranda_nilai WHERE id_indikator=:id AND jenis='realisasi' AND status_verifikasi='DISETUJUI' AND (nilai IS NOT NULL OR nilai_teks IS NOT NULL)",{"id":item["id_indikator"]})
            available=bool(latest and latest["jumlah"])
            verified_at=submission["diverifikasi_pada"] if submission else latest["waktu"] if available else None
            updater=submission["pengusul"] if submission and submission["pengusul"] else "Admin Provinsi" if available else None
            updater_role=submission["peran"] if submission else "ADMIN"
        else:
            available=bool(submission);verified_at=submission["diverifikasi_pada"] if submission else None;updater=submission["pengusul"] if submission else None;updater_role=submission["peran"] if submission else None
        status="Proxy" if item["is_proxy"] and available else "Tersedia" if available else "Belum Tersedia"
        evidence=rows(db,"SELECT id,nama_file,mime_type,ukuran,diunggah_pada FROM bukti_dukung WHERE usulan_id=:id ORDER BY id",{"id":submission["usulan_id"]}) if submission and submission.get("usulan_id") else []
        can_view=user["peran"] in {"ADMIN","VERIFIKATOR"} or user["peran"]=="OPERATOR" and submission and one(db,"SELECT 1 ok FROM usulan_nilai WHERE id=:id AND pengusul_id=:u",{"id":submission["usulan_id"],"u":user["id"]})
        metadata=one(db,"SELECT definisi,rumus_mentah,interpretasi,sumber_data,frekuensi FROM beranda_metadata WHERE id_indikator=:id",{"id":item["id_indikator"]})
        metadata_available=bool(metadata and any(metadata.values()))
        data.append({"id_indikator":item["id_indikator"],"kode_indikator":item["kode_indikator"],"nama_indikator":item["nama_indikator"],"satuan":item["satuan"],"instansi_pengampu":item["opd_pengampu"] or "Belum ditetapkan","validasi":f"Terverifikasi tanggal {verified_at}" if verified_at else "Belum diverifikasi","terverifikasi_pada":verified_at,"update":f"Terakhir update tanggal {verified_at} oleh {updater}" if verified_at and updater else "Belum ada pembaruan","update_oleh":updater,"peran_update":updater_role,"status_indikator":status,"metadata_tersedia":metadata_available,"usulan_id":submission["usulan_id"] if submission else None,"bukti_dukung_jumlah":len(evidence),"bukti_dukung":evidence if can_view else []})
    return {"wilayah":region,"wilayah_opsi":rows(db,"SELECT kode,nama,tingkat FROM wilayah WHERE aktif=1 ORDER BY length(kode),kode"),"data":data,"total":len(data),"status_data":"HANYA_TERVERIFIKASI"}


@router.get("/beranda-indikator/{indicator_id}/metadata")
def master_indicator_metadata(indicator_id:str,db:Session=Depends(get_db)):
    indicator=one(db,"SELECT id_indikator,kategori,kode_indikator,nama_indikator,kelompok,arah_pembangunan,satuan,opd_pengampu,status_ketersediaan,periode_data FROM beranda_indikator WHERE id_indikator=:id AND status_verifikasi='DISETUJUI'",{"id":indicator_id})
    if not indicator: raise HTTPException(404,"Indikator tidak ditemukan")
    metadata=one(db,"SELECT definisi,rumus_mentah,rumus_latex,interpretasi,sumber_data,frekuensi,status_metadata,sumber_metadata FROM beranda_metadata WHERE id_indikator=:id",{"id":indicator_id})
    values=rows(db,"SELECT tahun,jenis,nilai,nilai_teks,satuan_catatan FROM beranda_nilai WHERE id_indikator=:id AND status_verifikasi='DISETUJUI' ORDER BY tahun,CASE jenis WHEN 'realisasi' THEN 0 ELSE 1 END",{"id":indicator_id})
    return {**indicator,"metadata":metadata,"metadata_tersedia":bool(metadata and any(metadata.get(k) for k in ("definisi","rumus_mentah","interpretasi","sumber_data","frekuensi"))),"nilai":values}


def indicator_payload(db, ind, wilayah_kode=None):
    if wilayah_kode:
        vals=rows(db,"SELECT tahun,jenis,nilai,sumber sumber_sheet FROM nilai_indikator_wilayah WHERE id_indikator=:id AND wilayah_kode=:w ORDER BY tahun,jenis",{"id":ind["id_indikator"],"w":wilayah_kode})
    else:
        vals=rows(db,"SELECT tahun,jenis,nilai,sumber_sheet FROM nilai_indikator WHERE id_indikator=:id ORDER BY tahun,jenis",{"id":ind["id_indikator"]})
    actuals=[v for v in vals if v["jenis"]=="realisasi" and v["nilai"] is not None]
    last=max(actuals,key=lambda v:v["tahun"]) if actuals else None
    same=next((v for v in vals if last and v["jenis"]=="target" and v["tahun"]==last["tahun"] and v["nilai"] is not None),None)
    pct,status=achievement(last["nilai"] if last else None,same["nilai"] if same else None,ind.get("arah_baik"),ind.get("arah_baik_terverifikasi"))
    return {**ind,"nilai_terakhir":last["nilai"] if last else None,"tahun_terakhir_realisasi":last["tahun"] if last else None,"target_tahun_sama":same["nilai"] if same else None,"persentase_capaian":pct,"status_capaian":status,"tren":[{"tahun":v["tahun"],"nilai":v["nilai"]} for v in actuals]}


@router.post("/auth/login")
def login(form: OAuth2PasswordRequestForm=Depends(), db: Session=Depends(get_db)):
    user=one(db,"SELECT * FROM pengguna WHERE username=:u AND aktif=1",{"u":form.username})
    if not user or not password_hash.verify(form.password,user["password_hash"]): raise HTTPException(401,"Username atau kata sandi salah")
    token=jwt.encode({"sub":str(user["id"]),"exp":datetime.now(timezone.utc)+timedelta(hours=8)},SECRET,algorithm="HS256")
    return {"access_token":token,"token_type":"bearer","peran":user["peran"],"harus_ganti_password":bool(user["harus_ganti_password"])}


@router.get("/auth/saya")
def me(user=Depends(current_user)): return user


@router.post("/auth/ganti-password")
def change_password(password_baru:str=Form(...),user=Depends(require("ADMIN","OPERATOR","VERIFIKATOR")),db:Session=Depends(get_db)):
    if len(password_baru)<12: raise HTTPException(422,"Kata sandi minimal 12 karakter")
    db.execute(text("UPDATE pengguna SET password_hash=:p,harus_ganti_password=0 WHERE id=:id"),{"p":password_hash.hash(password_baru),"id":user["id"]});db.commit()
    return {"status":"PASSWORD_DIUBAH"}


@router.post("/admin/pengguna")
def create_user(username:str=Form(...),nama:str=Form(...),password:str=Form(...),peran:str=Form(...),wilayah_kode:str|None=Form(None),tim_pjk:str|None=Form(None),admin=Depends(require("ADMIN")),db:Session=Depends(get_db)):
    if peran not in {"ADMIN","OPERATOR","VERIFIKATOR","PENGUNJUNG"}:raise HTTPException(422,"Peran tidak valid")
    if peran in {"OPERATOR","VERIFIKATOR"} and not one(db,"SELECT kode FROM wilayah WHERE kode=:k AND aktif=1",{"k":wilayah_kode}):raise HTTPException(422,"Wilayah wajib dan harus aktif")
    if peran=="VERIFIKATOR" and wilayah_kode!="65":raise HTTPException(422,"Verifikator hanya dapat ditempatkan pada Provinsi Kalimantan Utara")
    if len(password)<12:raise HTTPException(422,"Kata sandi minimal 12 karakter")
    try:db.execute(text("INSERT INTO pengguna(username,nama,password_hash,peran,tim_pjk,wilayah_kode,harus_ganti_password) VALUES (:u,:n,:p,:r,:t,:w,1)"),{"u":username,"n":nama,"p":password_hash.hash(password),"r":peran,"t":tim_pjk,"w":wilayah_kode});db.commit()
    except Exception:db.rollback();raise HTTPException(409,"Username sudah digunakan")
    return {"status":"DIBUAT","username":username}


@router.get("/wilayah")
def regions(db:Session=Depends(get_db)):
    return {"data":rows(db,"SELECT kode,nama,tingkat,parent_kode FROM wilayah WHERE aktif=1 ORDER BY length(kode),kode")}


@router.get("/admin/pengguna")
def list_users(admin=Depends(require("ADMIN")),db:Session=Depends(get_db)):
    return {"data":rows(db,"SELECT p.id,p.username,p.nama,p.peran,p.wilayah_kode,w.nama wilayah,p.tim_pjk,p.aktif,p.harus_ganti_password FROM pengguna p LEFT JOIN wilayah w ON w.kode=p.wilayah_kode ORDER BY p.peran,w.kode,p.username")}


@router.patch("/admin/pengguna/{user_id}/status")
def set_user_status(user_id:int,aktif:bool=Form(...),admin=Depends(require("ADMIN")),db:Session=Depends(get_db)):
    if user_id==admin["id"] and not aktif: raise HTTPException(422,"Admin tidak dapat menonaktifkan akunnya sendiri")
    if not one(db,"SELECT id FROM pengguna WHERE id=:id",{"id":user_id}): raise HTTPException(404,"Pengguna tidak ditemukan")
    db.execute(text("UPDATE pengguna SET aktif=:a WHERE id=:id"),{"a":int(aktif),"id":user_id})
    db.execute(text("INSERT INTO log_aktivitas(pengguna_id,aksi,objek_tipe,objek_id,detail) VALUES (:u,'UBAH_STATUS_AKUN','pengguna',:id,:d)"),{"u":admin["id"],"id":str(user_id),"d":json.dumps({"aktif":aktif})});db.commit()
    return {"status":"AKTIF" if aktif else "NONAKTIF"}


@router.post("/admin/pengguna/{user_id}/reset-password")
def reset_user_password(user_id:int,password_baru:str=Form(...),admin=Depends(require("ADMIN")),db:Session=Depends(get_db)):
    if len(password_baru)<12:raise HTTPException(422,"Kata sandi minimal 12 karakter")
    if not one(db,"SELECT id FROM pengguna WHERE id=:id",{"id":user_id}):raise HTTPException(404,"Pengguna tidak ditemukan")
    db.execute(text("UPDATE pengguna SET password_hash=:p,harus_ganti_password=1 WHERE id=:id"),{"p":password_hash.hash(password_baru),"id":user_id})
    db.execute(text("INSERT INTO log_aktivitas(pengguna_id,aksi,objek_tipe,objek_id,detail) VALUES (:u,'RESET_PASSWORD','pengguna',:id,'Kata sandi direset oleh admin')"),{"u":admin["id"],"id":str(user_id)});db.commit()
    return {"status":"PASSWORD_DIRESET"}


@router.put("/arah-baik/{indicator_id}")
def correct_direction(indicator_id:str, arah_baik:str=Form(...), user=Depends(require("ADMIN")), db:Session=Depends(get_db)):
    if arah_baik not in {"NAIK","TURUN"}: raise HTTPException(422,"Arah harus NAIK atau TURUN")
    old=one(db,"SELECT arah_baik FROM indikator WHERE id_indikator=:id",{"id":indicator_id})
    if not old: raise HTTPException(404,"Indikator tidak ditemukan")
    db.execute(text("UPDATE indikator SET arah_baik=:a,arah_baik_terverifikasi=1 WHERE id_indikator=:id"),{"a":arah_baik,"id":indicator_id})
    db.execute(text("INSERT INTO log_perubahan(pengguna_id,id_indikator,field,nilai_lama,nilai_baru,sumber_perubahan) VALUES (:u,:id,'arah_baik',:o,:n,'koreksi_admin')"),{"u":user["id"],"id":indicator_id,"o":old["arah_baik"],"n":arah_baik});db.commit()
    return {"status":"ok","id_indikator":indicator_id,"arah_baik":arah_baik}


@router.get("/capaian")
def achievements(kategori:str|None=None,kelompok:str|None=None,arah_pembangunan:str|None=None,tim:str|None=None,status_capaian:str|None=None,wilayah_kode:str|None=None,db:Session=Depends(get_db)):
    if wilayah_kode and not one(db,"SELECT 1 ok FROM wilayah WHERE kode=:k AND aktif=1",{"k":wilayah_kode}):raise HTTPException(422,"Wilayah tidak valid")
    data=[indicator_payload(db,x,wilayah_kode) for x in rows(db,"SELECT id_indikator,nama_indikator,kategori,kelompok,arah_pembangunan,tim_pjk,satuan,arah_baik,arah_baik_terverifikasi FROM indikator ORDER BY kategori DESC,nomor")]
    for key,value in (("kategori",kategori),("kelompok",kelompok),("arah_pembangunan",arah_pembangunan),("tim_pjk",tim),("status_capaian",status_capaian)):
        if value: data=[x for x in data if x.get(key)==value]
    return {"data":data,"total":len(data),"arah_bersifat_sementara":True}


@router.get("/indikator/{indicator_id}/detail")
def detail(indicator_id:str,db:Session=Depends(get_db)):
    ind=one(db,"SELECT * FROM indikator WHERE id_indikator=:id",{"id":indicator_id})
    if not ind: raise HTTPException(404,"Indikator tidak ditemukan")
    payload=indicator_payload(db,ind)
    payload["nilai"]=rows(db,"SELECT tahun,jenis,nilai,sumber_sheet FROM nilai_indikator WHERE id_indikator=:id ORDER BY tahun,jenis",{"id":indicator_id})
    payload["metadata"]=one(db,"SELECT definisi,rumus_mentah,interpretasi,sumber_data,frekuensi,halaman_sumber,sumber_metadata,perlu_verifikasi_manual FROM metadata_indikator WHERE id_indikator=:id",{"id":indicator_id})
    return payload


@router.get("/indikator/{indicator_id}/unduh.csv")
def indicator_csv(indicator_id:str,db:Session=Depends(get_db)):
    ind=one(db,"SELECT nama_indikator FROM indikator WHERE id_indikator=:id",{"id":indicator_id})
    if not ind: raise HTTPException(404,"Indikator tidak ditemukan")
    output=StringIO();writer=csv.writer(output);writer.writerow(["ID Indikator","Nama Indikator","Tahun","Jenis","Nilai","Sumber"])
    for v in rows(db,"SELECT tahun,jenis,nilai,sumber_sheet FROM nilai_indikator WHERE id_indikator=:id ORDER BY tahun,jenis",{"id":indicator_id}):writer.writerow([indicator_id,ind["nama_indikator"],v["tahun"],v["jenis"],v["nilai"],v["sumber_sheet"]])
    return Response("\ufeff"+output.getvalue(),media_type="text/csv",headers={"Content-Disposition":f"attachment; filename={indicator_id}.csv"})


@router.get("/analitik/selisih/{indicator_id}")
def yearly_change(indicator_id:str,db:Session=Depends(get_db)):
    ind=one(db,"SELECT arah_baik,arah_baik_terverifikasi FROM indikator WHERE id_indikator=:id",{"id":indicator_id})
    vals=rows(db,"SELECT tahun,nilai FROM nilai_indikator WHERE id_indikator=:id AND jenis='realisasi' AND nilai IS NOT NULL ORDER BY tahun",{"id":indicator_id})
    data=[]
    for a,b in zip(vals,vals[1:]):
        diff=b["nilai"]-a["nilai"]; improvement=diff if ind and ind["arah_baik"]=="NAIK" else -diff
        data.append({"tahun":b["tahun"],"selisih":diff,"membaik":improvement>=0})
    return {"id_indikator":indicator_id,"arah_baik":ind["arah_baik"] if ind else None,"data":data}


@router.get("/analitik/peringkat")
def ranking(db:Session=Depends(get_db)):
    results=[]
    for ind in rows(db,"SELECT id_indikator,nama_indikator,arah_baik FROM indikator WHERE arah_baik_terverifikasi=1"):
        vals=rows(db,"SELECT tahun,nilai FROM nilai_indikator WHERE id_indikator=:id AND jenis='realisasi' AND nilai IS NOT NULL ORDER BY tahun DESC LIMIT 2",{"id":ind["id_indikator"]})
        if len(vals)==2:
            raw=vals[0]["nilai"]-vals[1]["nilai"];score=raw if ind["arah_baik"]=="NAIK" else -raw
            results.append({**ind,"tahun_awal":vals[1]["tahun"],"tahun_akhir":vals[0]["tahun"],"perubahan":raw,"skor_perbaikan":score})
    results.sort(key=lambda x:x["skor_perbaikan"],reverse=True)
    return {"perbaikan_terbesar":results[:10],"pemburukan_terbesar":list(reversed(results[-10:]))}


@router.get("/analitik/gap/{indicator_id}")
def gaps(indicator_id:str,db:Session=Depends(get_db)):
    ind=one(db,"SELECT nama_indikator,arah_baik,arah_baik_terverifikasi FROM indikator WHERE id_indikator=:id",{"id":indicator_id})
    if not ind: raise HTTPException(404,"Indikator tidak ditemukan")
    actual=rows(db,"SELECT tahun,nilai FROM nilai_indikator WHERE id_indikator=:id AND jenis='realisasi' AND nilai IS NOT NULL ORDER BY tahun",{"id":indicator_id})
    targets={x["tahun"]:x["nilai"] for x in rows(db,"SELECT tahun,nilai FROM nilai_indikator WHERE id_indikator=:id AND jenis='target' AND tahun IN (2029,2045)",{"id":indicator_id})}
    if not actual:return {"status":"BELUM_ADA_DATA","disclaimer":"Ekstrapolasi linear sederhana, bukan proyeksi resmi."}
    last=actual[-1]; hist=(last["nilai"]-actual[0]["nilai"])/(last["tahun"]-actual[0]["tahun"]) if len(actual)>1 and last["tahun"]!=actual[0]["tahun"] else None
    target=targets.get(2045); required=(target-last["nilai"])/(2045-last["tahun"]) if target is not None and last["tahun"]<2045 else None
    on_track=None
    if hist is not None and required is not None and ind["arah_baik_terverifikasi"]:on_track=hist>=required if ind["arah_baik"]=="NAIK" else hist<=required
    return {"id_indikator":indicator_id,"realisasi_terakhir":last,"target_2029":targets.get(2029),"target_2045":target,"gap_2029":targets.get(2029)-last["nilai"] if targets.get(2029)is not None else None,"gap_2045":target-last["nilai"] if target is not None else None,"laju_historis":hist,"required_run_rate":required,"status_jalur":"DI_JALUR" if on_track else "PERLU_AKSELERASI" if on_track is False else "BELUM_ADA_DATA","disclaimer":"Ekstrapolasi linear sederhana, bukan proyeksi resmi."}


@router.get("/analitik/multi")
def multi(ids:list[str]=Query(...),db:Session=Depends(get_db)):
    if len(ids)>4:raise HTTPException(422,"Maksimal empat indikator")
    return {"data":[{"id_indikator":iid,"nama":one(db,"SELECT nama_indikator FROM indikator WHERE id_indikator=:id",{"id":iid})["nama_indikator"],"seri":rows(db,"SELECT tahun,jenis,nilai FROM nilai_indikator WHERE id_indikator=:id AND nilai IS NOT NULL ORDER BY tahun",{"id":iid})} for iid in ids]}


@router.get("/analitik/korelasi")
def correlation(x:str,y:str,db:Session=Depends(get_db)):
    xv={r["tahun"]:r["nilai"] for r in rows(db,"SELECT tahun,nilai FROM nilai_indikator WHERE id_indikator=:id AND jenis='realisasi' AND nilai IS NOT NULL",{"id":x})};yv={r["tahun"]:r["nilai"] for r in rows(db,"SELECT tahun,nilai FROM nilai_indikator WHERE id_indikator=:id AND jenis='realisasi' AND nilai IS NOT NULL",{"id":y})};years=sorted(set(xv)&set(yv));points=[{"tahun":t,"x":xv[t],"y":yv[t]} for t in years]
    if len(points)<4:return {"n":len(points),"pearson":None,"data":points,"peringatan":"Hasil disembunyikan karena n < 4. Korelasi bukan sebab-akibat; seri pendek tidak layak ditafsirkan."}
    xs=[p["x"] for p in points];ys=[p["y"] for p in points];mx=sum(xs)/len(xs);my=sum(ys)/len(ys);den=math.sqrt(sum((a-mx)**2 for a in xs)*sum((b-my)**2 for b in ys));pearson=sum((a-mx)*(b-my) for a,b in zip(xs,ys))/den if den else None
    return {"n":len(points),"pearson":round(pearson,4) if pearson is not None else None,"data":points,"peringatan":"Korelasi bukan sebab-akibat; seri tahunan pendek harus ditafsirkan dengan sangat hati-hati."}


@router.post("/admin/usulan")
async def submit_value(id_indikator:str=Form(...),tahun:int=Form(...),jenis:str=Form(...),nilai:float=Form(...),periode:int|None=Form(None),sumber:str=Form(...),catatan:str|None=Form(None),wilayah_kode:str|None=Form(None),bukti:list[UploadFile]|None=File(None),user=Depends(require("ADMIN","OPERATOR")),db:Session=Depends(get_db)):
    if jenis not in {"realisasi","target"}:raise HTTPException(422,"Jenis tidak valid")
    if user["peran"]=="OPERATOR" and jenis!="realisasi":raise HTTPException(403,"Operator hanya dapat mengusulkan nilai realisasi")
    if not one(db,"SELECT 1 ok FROM indikator WHERE id_indikator=:id UNION SELECT 1 ok FROM beranda_indikator WHERE id_indikator=:id",{"id":id_indikator}):raise HTTPException(404,"Indikator tidak ditemukan")
    scope=user.get("wilayah_kode") if user["peran"]=="OPERATOR" else wilayah_kode
    if not scope or not one(db,"SELECT 1 ok FROM wilayah WHERE kode=:k AND aktif=1",{"k":scope}):raise HTTPException(422,"Wilayah tidak valid")
    attachments=bukti or []
    if not attachments: raise HTTPException(422,"Minimal satu bukti dukung wajib diunggah")
    allowed={"application/pdf","image/jpeg","image/png","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    prepared=[]
    for upload in attachments:
        content=await upload.read()
        if upload.content_type not in allowed: raise HTTPException(422,f"Format bukti tidak didukung: {upload.filename}")
        if len(content)>10*1024*1024: raise HTTPException(413,f"Bukti melebihi 10 MB: {upload.filename}")
        prepared.append((upload,content))
    if periode not in (None,1,2): raise HTTPException(422,"Periode semester harus 1 atau 2")
    result=db.execute(text("INSERT INTO usulan_nilai(id_indikator,tahun,jenis,nilai,periode,sumber,catatan,pengusul_id,wilayah_kode,dikirim_pada) VALUES (:i,:t,:j,:n,:p,:s,:c,:u,:w,CURRENT_TIMESTAMP) RETURNING id"),{"i":id_indikator,"t":tahun,"j":jenis,"n":nilai,"p":periode,"s":sumber,"c":catatan,"u":user["id"],"w":scope})
    submission_id=result.scalar_one(); target=EVIDENCE_DIR/str(submission_id);target.mkdir(parents=True,exist_ok=True)
    for upload,content in prepared:
        safe_name=f"{uuid.uuid4().hex}-{Path(upload.filename or 'bukti').name}";path=target/safe_name;path.write_bytes(content)
        db.execute(text("INSERT INTO bukti_dukung(usulan_id,nama_file,path_file,mime_type,ukuran,checksum_sha256) VALUES (:u,:n,:p,:m,:s,:h)"),{"u":submission_id,"n":upload.filename,"p":str(path),"m":upload.content_type,"s":len(content),"h":sha256(content).hexdigest()})
    db.execute(text("INSERT INTO log_aktivitas(pengguna_id,aksi,objek_tipe,objek_id,detail) VALUES (:u,'KIRIM_USULAN','usulan_nilai',:id,:d)"),{"u":user["id"],"id":str(submission_id),"d":json.dumps({"indikator":id_indikator,"tahun":tahun,"jenis":jenis,"wilayah":scope,"jumlah_bukti":len(prepared)})})
    db.commit();return {"status":"MENUNGGU_VERIFIKASI","id":submission_id,"jumlah_bukti":len(prepared)}


@router.get("/admin/usulan")
def submissions(status:str|None=None,user=Depends(require("ADMIN","OPERATOR","VERIFIKATOR")),db:Session=Depends(get_db)):
    where=[];params={}
    if status: where.append("u.status=:status");params["status"]=status
    if user["peran"]=="OPERATOR": where.append("u.pengusul_id=:uid");params["uid"]=user["id"]
    elif user["peran"]=="VERIFIKATOR" and user["wilayah_kode"]!="65": where.append("1=0")
    clause=" WHERE "+" AND ".join(where) if where else ""
    data=rows(db,"SELECT u.*,p.nama pengusul,p.peran peran_pengusul,w.nama wilayah,v.nama verifikator,(SELECT COUNT(*) FROM bukti_dukung b WHERE b.usulan_id=u.id) jumlah_bukti FROM usulan_nilai u JOIN pengguna p ON p.id=u.pengusul_id LEFT JOIN pengguna v ON v.id=u.verifikator_id LEFT JOIN wilayah w ON w.kode=u.wilayah_kode"+clause+" ORDER BY u.dibuat_pada DESC",params)
    return {"data":data}


def _submission_access(db,user,submission_id):
    item=one(db,"SELECT id,pengusul_id FROM usulan_nilai WHERE id=:id",{"id":submission_id})
    if not item:raise HTTPException(404,"Usulan tidak ditemukan")
    if user["peran"]=="OPERATOR" and item["pengusul_id"]!=user["id"]:raise HTTPException(403,"Bukti bukan milik usulan Anda")
    if user["peran"] not in {"ADMIN","VERIFIKATOR","OPERATOR"}:raise HTTPException(403,"Akses tidak diizinkan")
    return item


@router.get("/admin/usulan/{submission_id}/bukti")
def submission_evidence(submission_id:int,user=Depends(require("ADMIN","OPERATOR","VERIFIKATOR")),db:Session=Depends(get_db)):
    _submission_access(db,user,submission_id)
    return {"data":rows(db,"SELECT id,nama_file,mime_type,ukuran,checksum_sha256,diunggah_pada FROM bukti_dukung WHERE usulan_id=:id ORDER BY id",{"id":submission_id})}


@router.get("/admin/usulan/{submission_id}/bukti/{evidence_id}")
def view_evidence(submission_id:int,evidence_id:int,user=Depends(require("ADMIN","OPERATOR","VERIFIKATOR")),db:Session=Depends(get_db)):
    _submission_access(db,user,submission_id)
    evidence=one(db,"SELECT nama_file,path_file,mime_type FROM bukti_dukung WHERE id=:e AND usulan_id=:u",{"e":evidence_id,"u":submission_id})
    if not evidence:raise HTTPException(404,"Bukti dukung tidak ditemukan")
    path=Path(evidence["path_file"])
    if not path.exists():raise HTTPException(410,"File bukti dukung tidak tersedia di penyimpanan")
    return FileResponse(path,media_type=evidence["mime_type"] or "application/octet-stream",filename=evidence["nama_file"],content_disposition_type="inline")


@router.post("/admin/usulan/{submission_id}/verifikasi")
def verify_submission(submission_id:int,keputusan:str=Form(...),alasan:str|None=Form(None),user=Depends(require("ADMIN","VERIFIKATOR")),db:Session=Depends(get_db)):
    if keputusan not in {"DISETUJUI","DITOLAK"}:raise HTTPException(422,"Keputusan tidak valid")
    item=one(db,"SELECT * FROM usulan_nilai WHERE id=:id AND status='MENUNGGU_VERIFIKASI'",{"id":submission_id})
    if not item:raise HTTPException(404,"Usulan tidak ditemukan")
    if item["pengusul_id"]==user["id"]:raise HTTPException(403,"Pengusul tidak boleh memverifikasi usulannya sendiri")
    if user["peran"]=="VERIFIKATOR" and user["wilayah_kode"]!="65":raise HTTPException(403,"Verifikator harus bertugas di tingkat provinsi")
    if keputusan=="DITOLAK" and not alasan:raise HTTPException(422,"Alasan wajib untuk penolakan")
    if keputusan=="DISETUJUI":
        is_master=one(db,"SELECT 1 ok FROM beranda_indikator WHERE id_indikator=:i",{"i":item["id_indikator"]})
        old=one(db,"SELECT nilai FROM beranda_nilai_wilayah WHERE id_indikator=:i AND wilayah_kode=:w AND tahun=:t AND jenis=:j",{"i":item["id_indikator"],"w":item["wilayah_kode"],"t":item["tahun"],"j":item["jenis"]}) if is_master else one(db,"SELECT nilai FROM nilai_indikator_wilayah WHERE id_indikator=:i AND wilayah_kode=:w AND tahun=:t AND jenis=:j",{"i":item["id_indikator"],"w":item["wilayah_kode"],"t":item["tahun"],"j":item["jenis"]})
        published_value=item["nilai"]
        if is_master and item.get("periode"):
            db.execute(text("INSERT INTO beranda_nilai_wilayah_periode(id_indikator,wilayah_kode,tahun,jenis,periode,nilai,label_periode,sumber,usulan_id) VALUES (:i,:w,:t,:j,:p,:n,:l,:s,:u) ON CONFLICT(id_indikator,wilayah_kode,tahun,jenis,periode) DO UPDATE SET nilai=excluded.nilai,sumber=excluded.sumber,usulan_id=excluded.usulan_id,status_verifikasi='DISETUJUI',diverifikasi_pada=CURRENT_TIMESTAMP"),{"i":item["id_indikator"],"w":item["wilayah_kode"],"t":item["tahun"],"j":item["jenis"],"p":item["periode"],"n":item["nilai"],"l":f"Semester {item['periode']}","s":item["sumber"],"u":submission_id})
            latest=one(db,"SELECT nilai FROM beranda_nilai_wilayah_periode WHERE id_indikator=:i AND wilayah_kode=:w AND tahun=:t AND jenis=:j ORDER BY periode DESC LIMIT 1",{"i":item["id_indikator"],"w":item["wilayah_kode"],"t":item["tahun"],"j":item["jenis"]})
            published_value=latest["nilai"]
        if is_master: db.execute(text("INSERT INTO beranda_nilai_wilayah(id_indikator,wilayah_kode,tahun,jenis,nilai,sumber,usulan_id) VALUES (:i,:w,:t,:j,:n,:s,:u) ON CONFLICT(id_indikator,wilayah_kode,tahun,jenis) DO UPDATE SET nilai=excluded.nilai,sumber=excluded.sumber,usulan_id=excluded.usulan_id,status_verifikasi='DISETUJUI',diverifikasi_pada=CURRENT_TIMESTAMP"),{"i":item["id_indikator"],"w":item["wilayah_kode"],"t":item["tahun"],"j":item["jenis"],"n":published_value,"s":item["sumber"],"u":submission_id})
        else: db.execute(text("INSERT INTO nilai_indikator_wilayah(id_indikator,wilayah_kode,tahun,jenis,nilai,sumber,usulan_id) VALUES (:i,:w,:t,:j,:n,:s,:u) ON CONFLICT(id_indikator,wilayah_kode,tahun,jenis) DO UPDATE SET nilai=excluded.nilai,sumber=excluded.sumber,usulan_id=excluded.usulan_id,diverifikasi_pada=CURRENT_TIMESTAMP"),{"i":item["id_indikator"],"w":item["wilayah_kode"],"t":item["tahun"],"j":item["jenis"],"n":item["nilai"],"s":item["sumber"],"u":submission_id})
        if is_master and item["wilayah_kode"]=="65" and item.get("periode"): db.execute(text("INSERT INTO beranda_nilai_periode(id_indikator,tahun,jenis,periode,nilai,label_periode,sumber_master) VALUES (:i,:t,:j,:p,:n,:l,:s) ON CONFLICT(id_indikator,tahun,jenis,periode) DO UPDATE SET nilai=excluded.nilai,label_periode=excluded.label_periode,sumber_master=excluded.sumber_master,status_verifikasi='DISETUJUI',diverifikasi_pada=CURRENT_TIMESTAMP"),{"i":item["id_indikator"],"t":item["tahun"],"j":item["jenis"],"p":item["periode"],"n":item["nilai"],"l":f"Semester {item['periode']}","s":item["sumber"]})
        if is_master and item["wilayah_kode"]=="65":
            if item.get("periode"):
                latest=one(db,"SELECT nilai FROM beranda_nilai_periode WHERE id_indikator=:i AND tahun=:t AND jenis=:j ORDER BY periode DESC LIMIT 1",{"i":item["id_indikator"],"t":item["tahun"],"j":item["jenis"]})
                published_value=latest["nilai"]
            db.execute(text("INSERT INTO beranda_nilai(id_indikator,tahun,jenis,nilai,sumber_master,status_verifikasi) VALUES (:i,:t,:j,:n,:s,'DISETUJUI') ON CONFLICT(id_indikator,tahun,jenis) DO UPDATE SET nilai=excluded.nilai,nilai_teks=NULL,sumber_master=excluded.sumber_master,status_verifikasi='DISETUJUI',diverifikasi_pada=CURRENT_TIMESTAMP"),{"i":item["id_indikator"],"t":item["tahun"],"j":item["jenis"],"n":published_value,"s":item["sumber"]})
        # Preserve the legacy province dashboard while regional views use the scoped table.
        if item["wilayah_kode"]=="65": db.execute(text("INSERT INTO nilai_indikator(id_indikator,tahun,jenis,nilai,sumber_sheet) VALUES (:i,:t,:j,:n,:s) ON CONFLICT(id_indikator,tahun,jenis) DO UPDATE SET nilai=excluded.nilai,sumber_sheet=excluded.sumber_sheet"),{"i":item["id_indikator"],"t":item["tahun"],"j":item["jenis"],"n":published_value,"s":item["sumber"]})
        db.execute(text("INSERT INTO log_perubahan(pengguna_id,id_indikator,field,nilai_lama,nilai_baru,sumber_perubahan,referensi_id,catatan) VALUES (:u,:i,'nilai',:o,:n,'form',:r,:c)"),{"u":user["id"],"i":item["id_indikator"],"o":str(old["nilai"]) if old else None,"n":str(item["nilai"]),"r":str(submission_id),"c":item["catatan"]})
    db.execute(text("UPDATE usulan_nilai SET status=:k,verifikator_id=:u,diverifikasi_pada=CURRENT_TIMESTAMP,alasan_verifikasi=:a WHERE id=:id"),{"k":keputusan,"u":user["id"],"a":alasan,"id":submission_id})
    db.execute(text("INSERT INTO log_aktivitas(pengguna_id,aksi,objek_tipe,objek_id,detail) VALUES (:u,:aksi,'usulan_nilai',:id,:d)"),{"u":user["id"],"aksi":"SETUJUI_USULAN" if keputusan=="DISETUJUI" else "TOLAK_USULAN","id":str(submission_id),"d":json.dumps({"keputusan":keputusan,"alasan":alasan,"indikator":item["id_indikator"],"wilayah":item["wilayah_kode"]})});db.commit();return {"status":keputusan}


@router.get("/admin/log")
def audit_log(user=Depends(require("ADMIN")),db:Session=Depends(get_db)):return {"data":rows(db,"SELECT l.*,p.username FROM log_perubahan l LEFT JOIN pengguna p ON p.id=l.pengguna_id ORDER BY l.waktu DESC LIMIT 500")}


@router.post("/admin/unggah/pratinjau")
async def preview_upload(file:UploadFile=File(...),user=Depends(require("ADMIN")),db:Session=Depends(get_db)):
    if not file.filename.lower().endswith('.xlsx'):raise HTTPException(422,"Hanya file .xlsx")
    content=await file.read();
    if len(content)>30*1024*1024:raise HTTPException(413,"File melebihi 30 MB")
    ARCHIVE_DIR.mkdir(parents=True,exist_ok=True);uid=str(uuid.uuid4());archive=ARCHIVE_DIR/f"{datetime.now():%Y%m%d-%H%M%S}-{uid}.xlsx";archive.write_bytes(content)
    try:
        wb=load_workbook(archive,read_only=True);required={'form provinsi','ISV IUP Kaltara','ISV IUP Kaltara 2026','Rakor ISV IUP Kaltara 2026','Rakor ISV IUP Kaltara 202607'}
        missing=required-set(wb.sheetnames)
        if missing:raise HTTPException(422,f"Sheet hilang: {', '.join(sorted(missing))}")
        stage=archive.with_suffix(".stage.db");report=archive.with_suffix(".stage.md");run_etl(archive,stage,report)
        staged=sqlite3.connect(stage);current=sqlite3.connect(DEFAULT_DB)
        old={r[0]:r[1:] for r in current.execute("SELECT id_indikator,nama_indikator FROM indikator")};new={r[0]:r[1:] for r in staged.execute("SELECT id_indikator,nama_indikator FROM indikator")}
        oldv={(r[0],r[1],r[2]):r[3] for r in current.execute("SELECT id_indikator,tahun,jenis,nilai FROM nilai_indikator")};newv={(r[0],r[1],r[2]):r[3] for r in staged.execute("SELECT id_indikator,tahun,jenis,nilai FROM nilai_indikator")}
        diff={"indikator_baru":sorted(set(new)-set(old)),"indikator_hilang":sorted(set(old)-set(new)),"nilai_berubah":[{"id":k[0],"tahun":k[1],"jenis":k[2],"lama":oldv.get(k),"baru":newv.get(k)} for k in set(oldv)|set(newv) if oldv.get(k)!=newv.get(k)]}
        staged.close();current.close()
        result=db.execute(text("INSERT INTO unggahan_excel(nama_file_asli,path_arsip,checksum_sha256,status,ringkasan_diff,pengguna_id) VALUES (:n,:p,:h,'MENUNGGU_PERSETUJUAN',:d,:u) RETURNING id"),{"n":file.filename,"p":str(archive),"h":sha256(content).hexdigest(),"d":json.dumps(diff),"u":user["id"]}).scalar_one();db.commit();return {"id":result,"diff":diff}
    except HTTPException:raise
    except Exception as exc:raise HTTPException(422,f"ETL pratinjau gagal: {exc}")


@router.post("/admin/unggah/{upload_id}/setujui")
def approve_upload(upload_id:int,user=Depends(require("ADMIN")),db:Session=Depends(get_db)):
    item=one(db,"SELECT * FROM unggahan_excel WHERE id=:id AND status='MENUNGGU_PERSETUJUAN'",{"id":upload_id})
    if not item:raise HTTPException(404,"Unggahan tidak ditemukan")
    stage=Path(item["path_arsip"]).with_suffix(".stage.db")
    if not stage.exists():raise HTTPException(409,"Database staging tidak ditemukan")
    source=sqlite3.connect(stage);values=source.execute("SELECT id_indikator,tahun,jenis,nilai,sumber_sheet FROM nilai_indikator").fetchall();source.close()
    for iid,year,kind,value,source_sheet in values:
        old=one(db,"SELECT nilai FROM nilai_indikator WHERE id_indikator=:i AND tahun=:t AND jenis=:j",{"i":iid,"t":year,"j":kind})
        if not old or old["nilai"]!=value:
            db.execute(text("INSERT INTO nilai_indikator(id_indikator,tahun,jenis,nilai,sumber_sheet) VALUES (:i,:t,:j,:v,:s) ON CONFLICT(id_indikator,tahun,jenis) DO UPDATE SET nilai=excluded.nilai,sumber_sheet=excluded.sumber_sheet"),{"i":iid,"t":year,"j":kind,"v":value,"s":source_sheet})
            db.execute(text("INSERT INTO log_perubahan(pengguna_id,id_indikator,field,nilai_lama,nilai_baru,sumber_perubahan,referensi_id) VALUES (:u,:i,:f,:o,:n,'unggah',:r)"),{"u":user["id"],"i":iid,"f":f"nilai:{year}:{kind}","o":str(old["nilai"]) if old else None,"n":str(value),"r":str(upload_id)})
    db.execute(text("UPDATE unggahan_excel SET status='DISETUJUI',disetujui_pada=CURRENT_TIMESTAMP WHERE id=:id"),{"id":upload_id});db.commit();return {"status":"DISETUJUI"}


def metadata_workbook(db):
    wb=Workbook();ws=wb.active;ws.title="Katalog Metadata";headers=["ID Indikator","Nama Indikator","Definisi","Rumus Mentah","Interpretasi","Sumber Data","Frekuensi","Sumber Metadata"];ws.append(headers)
    for r in rows(db,"SELECT i.id_indikator,i.nama_indikator,m.definisi,m.rumus_mentah,m.interpretasi,m.sumber_data,m.frekuensi,m.sumber_metadata FROM indikator i LEFT JOIN metadata_indikator m USING(id_indikator) ORDER BY i.kategori DESC,i.nomor"):ws.append(list(r.values()))
    stream=BytesIO();wb.save(stream);return stream.getvalue()


def metadata_pdf(db):
    stream=BytesIO();c=canvas.Canvas(stream,pagesize=A4);w,h=A4;y=h-45;c.setFont("Helvetica-Bold",15);c.drawString(40,y,"Katalog Metadata Indikator SEBATIK");y-=28
    for r in rows(db,"SELECT i.id_indikator,i.nama_indikator,m.definisi FROM indikator i LEFT JOIN metadata_indikator m USING(id_indikator) ORDER BY i.kategori DESC,i.nomor"):
        if y<70:c.showPage();y=h-45
        c.setFont("Helvetica-Bold",8);c.drawString(40,y,f"{r['id_indikator']} - {r['nama_indikator'][:85]}");y-=11;c.setFont("Helvetica",7)
        text=(r["definisi"] or "Metadata belum tersedia")[:350]
        for start in range(0,len(text),105):c.drawString(45,y,text[start:start+105]);y-=9
        y-=5
    c.save();return stream.getvalue()


@router.get("/download/paket.zip")
def download_package(db:Session=Depends(get_db)):
    stream=BytesIO()
    with zipfile.ZipFile(stream,'w',zipfile.ZIP_DEFLATED) as z:
        for table in ('indikator','nilai_indikator','metadata_indikator'):
            data=rows(db,f"SELECT * FROM {table}");out=StringIO();writer=csv.DictWriter(out,fieldnames=list(data[0]) if data else []);writer.writeheader();writer.writerows(data);z.writestr(f"{table}.csv",'\ufeff'+out.getvalue())
        z.writestr("katalog-metadata.xlsx",metadata_workbook(db));z.writestr("katalog-metadata.pdf",metadata_pdf(db))
    stream.seek(0);return StreamingResponse(stream,media_type="application/zip",headers={"Content-Disposition":"attachment; filename=paket-data-sebatik.zip"})
