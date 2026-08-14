"""Pipeline ETL workbook ISV-IUP menjadi SQLite dan cadangan tabular."""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path
import re
import sqlite3
from typing import Any

import openpyxl

from .common import clean_text, enum_rpjmd, indicator_id, parse_angka
from .units import indicator_unit


VALUE_SOURCES = [
    "Rakor ISV IUP Kaltara 202607",
    "Rakor ISV IUP Kaltara 2026",
    "ISV IUP Kaltara 2026",
    "ISV IUP Kaltara",
]


def header_index(ws) -> dict[str, int]:
    return {clean_text(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1) if clean_text(ws.cell(1, c).value)}


def extract_unit(name: str | None) -> str | None:
    if not name:
        return None
    matches = re.findall(r"\(([^()]+)\)", name)
    if matches:
        candidate = matches[-1].strip()
        if len(candidate) <= 40:
            return candidate
    if "indeks" in name.casefold():
        return "indeks"
    if "rasio" in name.casefold():
        return "rasio"
    return None


def extract_proxy(flag: Any, note: Any) -> tuple[int, str | None]:
    f, n = clean_text(flag), clean_text(note)
    text = " ".join(x for x in (f, n) if x)
    is_proxy = bool(f and f.casefold() not in {"tidak", "tidak ada", "-"}) or "proxy" in text.casefold()
    match = re.search(r"indikator\s+proxy\s*:\s*([^;\n]+)", str(note or ""), flags=re.I)
    return int(is_proxy), clean_text(match.group(1)) if match else (f if is_proxy and f and f.casefold() not in {"ya", "proxy"} else None)


def master_rows(wb):
    ws = wb["form provinsi"]
    h = header_index(ws)
    def val(row, label):
        col = h.get(label)
        return ws.cell(row, col).value if col else None
    rows, pics = [], []
    sequence = {"ISV": 0, "IUP": 0}
    for r in range(2, 88):
        category = (clean_text(val(r, "Kategori")) or "").upper()
        if category not in sequence:
            continue
        sequence[category] += 1
        iid = indicator_id(category, sequence[category])
        if not iid:
            continue
        original = clean_text(val(r, "Indikator"))
        improved = clean_text(val(r, "Perbaikan Nama Indikator"))
        proxy, proxy_name = extract_proxy(val(r, "Indikator Proxy"), val(r, "Catatan Teknis"))
        year = parse_angka(val(r, "Tahun Terakhir Data"))
        rows.append({
            "id_indikator": iid, "kategori": category, "nomor": sequence[category],
            "nama_indikator": improved or original, "nama_asli": original,
            "kelompok": clean_text(val(r, "Kelompok Indikator")),
            "arah_pembangunan": clean_text(val(r, "Arah Pembangunan")),
            "satuan": indicator_unit(iid), "penghasil": clean_text(val(r, "Penghasil Indikator")),
            "kl_pengampu": clean_text(val(r, "K/L/D/i Pengampu Indikator")),
            "opd_penanggung_jawab": None, "tim_pjk": clean_text(val(r, "Tim PJK Provinsi")),
            "status_ketersediaan": clean_text(val(r, "Ketersediaan")),
            "status_metadata": clean_text(val(r, "Ketersediaan Metadata Indikator")) or "Tidak Tersedia",
            "periode_data": clean_text(val(r, "Periode Data")) or clean_text(val(r, "Periode Data ")),
            "tahun_terakhir": int(year) if year is not None and 1900 <= year <= 2100 else None,
            "is_proxy": proxy, "nama_proxy": proxy_name,
            "status_rpjmd": enum_rpjmd(val(r, "NOTES to DANS")),
            "kode_sdgs": clean_text(val(r, "Indikator SDGs")),
            "link_metadata": clean_text(val(r, "Tautan Metadata Indikator")),
            "link_publikasi": clean_text(val(r, "Link Publikasi")), "link_data": clean_text(val(r, "Link Data")),
            "catatan_teknis": clean_text(val(r, "Catatan Teknis")),
        })
        for kind, label in (("PIC_DGDS", "PIC DGDS"), ("PIC_PJK", "PIC PJK"), ("PIC_PROVINSI", "PIC Provinsi")):
            person = clean_text(val(r, label))
            if person:
                pics.append({"id_indikator": iid, "jenis_pic": kind, "nama_pic": person})
    return rows, pics


def keyed_old_rows(ws, name_col=2):
    for r in range(3, min(ws.max_row, 200) + 1):
        number, name = parse_angka(ws.cell(r, 1).value), clean_text(ws.cell(r, name_col).value)
        if number is None or not name:
            continue
        category = "ISV" if int(number) <= 10 else "IUP"
        iid = indicator_id(category, number if category == "ISV" else number - 10)
        if iid:
            yield r, iid, name


def enrich_owners(wb, indicators):
    ws = wb["ISV IUP Kaltara 2026"]
    owner = {iid: (clean_text(ws.cell(r, 3).value), clean_text(ws.cell(r, 4).value)) for r, iid, _ in keyed_old_rows(ws)}
    for row in indicators:
        opd, team = owner.get(row["id_indikator"], (None, None))
        row["opd_penanggung_jawab"] = opd
        # Tim PJK tetap berasal dari master `form provinsi`; sheet 2026 hanya
        # menjadi sumber OPD agar domain master (48/30/5/3) tidak berubah.


@dataclass
class ParseStats:
    success: int = 0
    failed: int = 0
    blank: int = 0


def add_value(store, stats, iid, year, kind, raw, source):
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        stats.blank += 1
        return
    value = parse_angka(raw)
    if value is None:
        stats.failed += 1
        return
    stats.success += 1
    store.setdefault((iid, int(year), kind), {"id_indikator": iid, "tahun": int(year), "jenis": kind, "nilai": value, "sumber_sheet": source})


def extract_values(wb):
    store, stats = {}, ParseStats()
    # Prioritas 1: format target/realisasi berpasangan.
    ws = wb[VALUE_SOURCES[0]]
    cat = num = None
    for r in range(3, 165):
        cat = ws.cell(r, 1).value or cat
        num = ws.cell(r, 2).value or num
        iid = indicator_id(cat, num if (clean_text(cat) or "").upper() == "ISV" else (parse_angka(num) or 0) - 10)
        kind = (clean_text(ws.cell(r, 4).value) or "").casefold()
        if iid and kind in {"target", "realisasi"}:
            for c, year in zip(range(5, 14), range(2021, 2030)):
                add_value(store, stats, iid, year, kind, ws.cell(r, c).value, ws.title)
            if kind == "target":
                add_value(store, stats, iid, 2045, kind, ws.cell(r, 14).value, ws.title)
    # Prioritas 2: Rakor 2026, kolom realisasi dan target.
    ws = wb[VALUE_SOURCES[1]]
    for r in range(3, 200):
        category, number = ws.cell(r,1).value, parse_angka(ws.cell(r,2).value)
        iid = indicator_id(category, number if (clean_text(category) or "").upper() == "ISV" else (number - 10 if number is not None else None))
        if not iid: continue
        for c, year in zip(range(6,11), range(2021,2026)):
            add_value(store, stats, iid, year, "realisasi", ws.cell(r,c).value, ws.title)
        for c, year in zip(range(11,16), range(2025,2030)):
            add_value(store, stats, iid, year, "target", ws.cell(r,c).value, ws.title)
    # Prioritas 3: realisasi 2026.
    ws = wb[VALUE_SOURCES[2]]
    for r, iid, _ in keyed_old_rows(ws):
        for c, year in zip(range(5,11), range(2021,2027)):
            add_value(store, stats, iid, year, "realisasi", ws.cell(r,c).value, ws.title)
    # Prioritas 4: format lama.
    ws = wb[VALUE_SOURCES[3]]
    for r, iid, _ in keyed_old_rows(ws):
        for c, year in zip(range(3,8), range(2021,2026)):
            add_value(store, stats, iid, year, "realisasi", ws.cell(r,c).value, ws.title)
        for c, year in zip(range(8,13), range(2025,2030)):
            add_value(store, stats, iid, year, "target", ws.cell(r,c).value, ws.title)
    return list(store.values()), stats


SCHEMA = """
PRAGMA foreign_keys=ON;
CREATE TABLE indikator (
 id_indikator TEXT PRIMARY KEY, kategori TEXT NOT NULL CHECK(kategori IN ('ISV','IUP')), nomor INTEGER NOT NULL,
 nama_indikator TEXT NOT NULL, nama_asli TEXT, kelompok TEXT, arah_pembangunan TEXT, satuan TEXT,
 penghasil TEXT, kl_pengampu TEXT, opd_penanggung_jawab TEXT, tim_pjk TEXT,
 status_ketersediaan TEXT, status_metadata TEXT, periode_data TEXT, tahun_terakhir INTEGER,
 is_proxy INTEGER NOT NULL DEFAULT 0 CHECK(is_proxy IN (0,1)), nama_proxy TEXT,
 status_rpjmd TEXT NOT NULL CHECK(status_rpjmd IN ('MASUK_RPJMD','TIDAK_MASUK_RPJMD','MASUK_TAPI_BELUM_ADA_DATA','DOBEL_ISV_IUP')),
 arah_baik TEXT CHECK(arah_baik IN ('NAIK','TURUN')),
 arah_baik_terverifikasi INTEGER NOT NULL DEFAULT 0 CHECK(arah_baik_terverifikasi IN (0,1)),
 kode_sdgs TEXT, link_metadata TEXT, link_publikasi TEXT, link_data TEXT, catatan_teknis TEXT,
 UNIQUE(kategori, nomor)
);
CREATE TABLE nilai_indikator (
 id_indikator TEXT NOT NULL REFERENCES indikator(id_indikator), tahun INTEGER NOT NULL,
 jenis TEXT NOT NULL CHECK(jenis IN ('realisasi','target')), nilai REAL, sumber_sheet TEXT NOT NULL,
 PRIMARY KEY(id_indikator,tahun,jenis)
);
CREATE TABLE metadata_indikator (
 id_indikator TEXT PRIMARY KEY REFERENCES indikator(id_indikator), definisi TEXT, rumus TEXT,
 rumus_mentah TEXT, interpretasi TEXT, sumber_data TEXT, frekuensi TEXT, halaman_sumber TEXT,
 perlu_verifikasi_manual INTEGER NOT NULL DEFAULT 0, sumber_metadata TEXT, nama_di_buku1 TEXT
);
CREATE TABLE penugasan_pic (
 id INTEGER PRIMARY KEY AUTOINCREMENT, id_indikator TEXT NOT NULL REFERENCES indikator(id_indikator),
 jenis_pic TEXT NOT NULL, nama_pic TEXT NOT NULL
);
CREATE TABLE snapshot_ketersediaan (
 id_indikator TEXT NOT NULL REFERENCES indikator(id_indikator), tanggal_snapshot TEXT NOT NULL,
 status TEXT NOT NULL, PRIMARY KEY(id_indikator,tanggal_snapshot)
);
"""


def write_csv(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows: return
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0]))
        writer.writeheader(); writer.writerows(rows)


def run(workbook_path: Path, db_path: Path, report_path: Path):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    indicators, pics = master_rows(wb)
    enrich_owners(wb, indicators)
    values, stats = extract_values(wb)
    if len(indicators) != 86:
        raise RuntimeError(f"Dimensi indikator harus 86, ditemukan {len(indicators)}")
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists(): db_path.unlink()
    conn = sqlite3.connect(db_path)
    conn.executescript(SCHEMA)
    ind_cols = list(indicators[0])
    conn.executemany(f"INSERT INTO indikator ({','.join(ind_cols)}) VALUES ({','.join('?' for _ in ind_cols)})", [[r[c] for c in ind_cols] for r in indicators])
    val_cols = list(values[0])
    conn.executemany(f"INSERT INTO nilai_indikator ({','.join(val_cols)}) VALUES ({','.join('?' for _ in val_cols)})", [[r[c] for c in val_cols] for r in values])
    conn.executemany("INSERT INTO metadata_indikator(id_indikator) VALUES (?)", [(r["id_indikator"],) for r in indicators])
    conn.executemany("INSERT INTO penugasan_pic(id_indikator,jenis_pic,nama_pic) VALUES (?,?,?)", [(r["id_indikator"],r["jenis_pic"],r["nama_pic"]) for r in pics])
    from datetime import date
    conn.executemany("INSERT INTO snapshot_ketersediaan(id_indikator,tanggal_snapshot,status) VALUES (?,?,?)", [(r["id_indikator"], date.today().isoformat(), r["status_ketersediaan"] or "Belum Tersedia") for r in indicators])
    conn.commit()
    fk_errors = conn.execute("PRAGMA foreign_key_check").fetchall()
    counts = {t: conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0] for t in ("indikator","nilai_indikator","metadata_indikator","penugasan_pic")}
    no_actual = [x[0] for x in conn.execute("SELECT i.id_indikator FROM indikator i LEFT JOIN nilai_indikator n ON n.id_indikator=i.id_indikator AND n.jenis='realisasi' AND n.nilai IS NOT NULL GROUP BY i.id_indikator HAVING COUNT(n.id_indikator)=0 ORDER BY i.id_indikator")]
    conn.close()
    backup = db_path.parent / "cadangan"
    write_csv(backup / "indikator.csv", indicators)
    write_csv(backup / "nilai_indikator.csv", values)
    write_csv(backup / "penugasan_pic.csv", pics)
    write_csv(backup / "metadata_indikator.csv", [{"id_indikator": r["id_indikator"]} for r in indicators])
    parquet_status = "tidak dibuat (pyarrow tidak tersedia)"
    try:
        import pandas as pd
        for name, rows in (("indikator",indicators),("nilai_indikator",values),("penugasan_pic",pics)):
            pd.DataFrame(rows).to_parquet(backup / f"{name}.parquet", index=False)
        parquet_status = "dibuat"
    except (ImportError, ModuleNotFoundError):
        pass
    report = ["# Laporan ETL SEBATIK", "", "## Ringkasan tabel", "", "| Tabel | Jumlah baris |", "|---|---:|"]
    report += [f"| {k} | {v} |" for k,v in counts.items()]
    report += ["", "## Validasi parsing nilai", "", f"- Berhasil di-parse: **{stats.success}** sel sumber.", f"- Gagal di-parse: **{stats.failed}** sel sumber nonkosong.", f"- Kosong dan dipertahankan sebagai NULL/tidak dibuat: **{stats.blank}** sel sumber.", f"- Fakta unik setelah prioritas sumber: **{len(values)}** baris.", f"- Cadangan Parquet: **{parquet_status}**; CSV selalu dibuat.", f"- Pelanggaran foreign key: **{len(fk_errors)}**.", "", f"## Indikator tanpa satu pun nilai realisasi ({len(no_actual)})", ""]
    report += [f"- {iid}" for iid in no_actual] or ["Tidak ada."]
    report += ["", "## Aturan provenans", "", "Urutan prioritas: `Rakor ISV IUP Kaltara 202607` -> `Rakor ISV IUP Kaltara 2026` -> `ISV IUP Kaltara 2026` -> `ISV IUP Kaltara`. Sumber lama hanya mengisi kombinasi indikator-tahun-jenis yang masih kosong."]
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(report)+"\n", encoding="utf-8")
    print(f"ETL selesai: {counts}; parse berhasil={stats.success}, gagal={stats.failed}")


if __name__ == "__main__":
    p = argparse.ArgumentParser(); p.add_argument("workbook", type=Path); p.add_argument("--db", type=Path, default=Path("data/processed/sebatik.db")); p.add_argument("--report", type=Path, default=Path("docs/02-etl-report.md")); a=p.parse_args()
    run(a.workbook, a.db, a.report)
